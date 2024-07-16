import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk, filedialog, messagebox
from combobox_options import (TYPE_OPTIONS, CLASS_OPTIONS, REPORTING_GROUP_OPTIONS,
                              ON_HOLD_REASON_OPTIONS, GROUP_OPTIONS, LABEL_GROUP_OPTIONS)
from openpyxl.utils import get_column_letter, exceptions, column_index_from_string
import openpyxl
import sys
import os
import re
import msvcrt
import gc


# region Validation Methods

def check_empty_rows(file_path, sheet_index, column, row_start, row_end):
    """
    Checks for empty rows in a specified column for a specified range of rows

    :param file_path: The file containing the sheet and column the user intends to check
    :param sheet_index: The index of the sheet within the file_path parameter
    :param column: The specified column
    :param row_start: The first row the program should check
    :param row_end: The last row the program should check
    :return: A list of the rows that were empty
    """

    workbook = None
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, keep_vba=False, data_only=True)

        if sheet_index < 0 or sheet_index >= len(workbook.sheetnames):
            raise ValueError(
                f"Sheet index {sheet_index} is out of range. The workbook has {len(workbook.sheetnames)} sheets.")

        sheet = workbook.worksheets[sheet_index]
        empty_rows = []

        # Convert column reference to index if it's a string
        column = column_index_from_string(column) if isinstance(column, str) else column

        # Ensure row_start and row_end are within the sheet's range
        max_row = sheet.max_row
        row_start = max(1, min(row_start, max_row))
        row_end = min(row_end, max_row)

        for row in range(row_start, row_end + 1):
            cell_value = sheet.cell(row=row, column=column).value

            if cell_value is None or cell_value == "":
                empty_rows.append(row)

        return empty_rows

    finally:
        if workbook:
            workbook.close()
        gc.collect()  # Force garbage collection


def sheet_exists(excel_file_path, sheet_name):
    """
    Check if a sheet exists in an Excel file.

    :param excel_file_path: The path to the Excel file
    :type excel_file_path: str
    :param sheet_name: The name of the sheet to check
    :type sheet_name: str

    :return: True if the sheet exists, False otherwise
    :rtype: bool
    """
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        return sheet_name in workbook.sheetnames
    except FileNotFoundError:
        print(f"File not found: {excel_file_path}")
        return False
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"Invalid Excel file: {excel_file_path}")
        return False
    except Exception(BaseException):
        print("An error occured. Please try again.")


def get_sheet_index(excel_file_path, sheet_name):
    """
    Get the index of a sheet in an Excel file.

    :param excel_file_path: The file path of the Excel file.
    :type excel_file_path: str
    :param sheet_name: The name of the sheet to find.
    :type sheet_name: str

    :return: The index of the sheet in the Excel file.
    :rtype: int
    """

    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet_index = workbook.sheetnames.index(sheet_name)
        return sheet_index
    except FileNotFoundError:
        print(f"File not found: {excel_file_path}")
        return False
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"Invalid Excel file: {excel_file_path}")
        return False
    except ValueError:
        print(f"Sheet '{sheet_name}' not found in the Excel file.")
        return False


def validate_file_location(file_path):
    """
    Validate the file location.

    :param file_path: The path to the file
    :type file_path: str

    :return: True if the file location is valid, False otherwise
    :rtype: bool
    """

    # Check if the file path is not empty
    if not file_path:
        return False

    # Check if the file path exists
    if not os.path.exists(file_path):
        return False

    # Check if the file path points to a file (not a directory)
    if os.path.isdir(file_path):
        return False

    return True


def is_file_open(file_path):
    try:
        # Try to open the file in read-write mode
        with open(file_path, 'r+b') as f:
            # Try to acquire a lock on the file
            msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
            # If we got here, the file wasn't locked
            msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
        return False
    except IOError:
        # If we got an IOError, the file is likely open by another process
        return True
    except WindowsError as e:
        # Windows-specific errors
        if e.winerror == 32:  # ERROR_SHARING_VIOLATION
            return True
        elif e.winerror == 33:  # ERROR_LOCK_VIOLATION
            return True
        else:
            raise  # Re-raise any other Windows errors


def is_valid_row_combo(first_row, last_row):
    """
    Checks if the inputs are valid rows and a valid row combination

    :param first_row: The first row of the part number list
    :type first_row: int
    :param last_row:  The last row of the part number list
    :type last_row: int

    :return: True if the rows are individually valid and a valid combination. False if any one of those is not true
    :rtype: bool
    """

    # Check if both are valid integers
    if is_valid_integer(first_row) and is_valid_integer(last_row):
        # Check if the first row is less than the last row
        if int(first_row) < int(last_row):
            # Check if more than 150 parts are being created
            if int(last_row) - int(first_row) < 150:
                return True
            else:
                return False
        else:
            return False
    else:
        return False


def is_valid_column(column):
    """
    Check if the input is a valid column

    :param column: A letter representing a column of an Excel spreadsheet
    :type column: str

    :return: True if the column letter is a letter or combination of letters. False if invalid
    """

    return re.match(r'^[A-Za-z]+$', column)


def is_valid_integer(var):
    """
    Checks if input is a valid integer

    :param var: A number representing a row in and Excel spreadsheet
    :type var: int

    :return: True if var is a valid integer. False if var is not a valid integer
    """

    return isinstance(var, str) and var.isdigit()


def browse_file(var):
    """
    Sets the selected file path to the given `var`.

    :param var: tkinter variable to store the selected file path
    """

    file_path = filedialog.askopenfilename()
    var.set(file_path)


# endregion


class BaseForm:
    def __init__(self, master):
        """
        Initialize the BaseForm class and the master tkinter form

        :param master: The master widget (usually a Tk instance) for the form.
        :type master: tk.Tk
        """
        self.master = master  # Set the master widget
        self.master.title("Create Part")  # Set the title of the form
        self.master.protocol("WM_DELETE_WINDOW", self.on_closing)  # Define the protocol for closing the window

        # Initialize miscellaneous variables
        self.file_data = {}  # Dictionary to store file-related data
        self.label_data = {}  # Dictionary to store label-related data
        self.file_widgets = []  # List to hold file-related widgets
        self.label_widgets = []  # List to hold label-related widgets
        self.small_font = tkfont.Font(size=12)  # Define a small font for the form
        self.is_terminated = False  # Flag to track if the form is terminated

    # region Widget Creation
    def create_entry_widget(self, frame, label, row, col, arr, var_type=tk.StringVar):
        """
        Creates a label and an adjacent textbox into the given frame at a specified row/column combination

        :param frame: The frame where controls are placed
        :param label: The text value of the label
        :param row: The row in which both the label and the textbox will be placed
        :param col: The column in which the label will be placed (the textbox is placed one column to the right)
        :param arr: The array where the widgets will be listed
        :param var_type: The type of variable that will be passed, in combination with the label, as the value to arr
        :return: None
        """

        label_widget = ttk.Label(frame, text=label, width=22, anchor='w', font=self.small_font)
        label_widget.grid(row=row, column=col, padx=(0, 10), pady=7, sticky='w')

        var = var_type()
        entry = ttk.Entry(frame, textvariable=var)
        entry.grid(row=row, column=col + 1, padx=(0, 10), pady=7, sticky='ew')

        arr.append((label, var))

    def create_file_widget(self, frame, label, row, col, arr, var_type=tk.StringVar):
        """
        Creates a label, an adjacent textbox, and an adjacent button into the given frame at a specified
        row/column combination. The button will allow users to enter File Explorer and the file they choose will
        become the value within the textbox.

        :param frame: The frame where controls are placed
        :param label: The text value of the label
        :param row: The row in which the label, textbox, and button will be placed
        :param col: The column in which the label will be placed (the textbox is placed one column to the right and the
        button is placed two columns to the right)
        :param arr: The array where the widgets will be listed
        :param var_type: The type of variable that will be passed, in combination with the label, as the value to arr
        :return: None
        """

        label_widget = ttk.Label(frame, text=label, width=20, anchor='w', font=self.small_font)
        label_widget.grid(row=row, column=col, padx=(0, 10), pady=7, sticky='w')

        var = var_type()
        entry = ttk.Entry(frame, textvariable=var)
        entry.grid(row=row, column=col + 1, padx=(0, 10), pady=7, sticky='ew')

        browse_button = tk.Button(frame, text="Browse", command=lambda: browse_file(var))
        browse_button.grid(row=row, column=col + 2, padx=(0, 10), pady=7, sticky='ew')

        arr.append((label, var))

    def create_dropdown_widget(self, frame, label, width, options, row, col, arr, var_type=tk.StringVar):
        """
        Creates a label and an adjacent dropdown menu into the given frame at a specified row/column combination

        :param frame: The frame where controls are placed
        :param label: The text value of the label
        :param width: The width of the dropdown menu
        :param options: The choices within the dropdown menu
        :param row: The row in which both the label and the dropdown menu will be placed
        :param col: The column in which the label will be placed (the dropdown menu is placed one column to the right)
        :param arr: The array where the widgets will be listed
        :param var_type: The type of variable that will be passed, in combination with the label, as the value to arr
        :return: None
        """

        label_widget = ttk.Label(frame, text=label, width=20, anchor='w', font=self.small_font)
        label_widget.grid(row=row, column=col, padx=(0, 10), pady=7, sticky='w')

        var = var_type()
        combobox = ttk.Combobox(frame, textvariable=var, values=options, state="readonly")
        combobox.grid(row=row, column=col + 1, padx=(0, 10), pady=7, sticky='w')
        combobox.config(width=width)

        arr.append((label, var))

    def create_checkbox_widget(self, frame, label, row, col, arr, var_type=tk.BooleanVar):
        """
        Creates a checkbox and assigns it to a specific row/col combination in the grid

        :param frame: The frame where controls are placed
        :param label: The label next to the checkbox
        :param row: The row in which the checkbox will be placed
        :param col: The column in which the checkbox will be placed
        :param arr: The array where the widgets will be listed
        :param var_type: :param var_type: The type of variable that will be passed, in combination with the label,
         as the value to arr
        :return:
        """
        var = var_type()  # Create a BooleanVar instance that updates when the checkbox state is changed

        checkbox = ttk.Checkbutton(frame, text=label, variable=var)
        checkbox.grid(row=row, column=col, padx=(0, 10), pady=7, sticky='w')

        arr.append((label, var))

    # endregion

    def on_closing(self):
        """
        Sends a messagebox to the user confirming the termination of the program

        :return: None
        """

        if messagebox.askokcancel("Quit", "Do you want to quit the program?"):
            self.is_terminated = True
            self.master.destroy()
            sys.exit()

    def create_file_form(self, operation_type):
        """
        This method should be overridden by subclasses to implement specific logic

        :param operation_type: The specific operation type (OperationType.Create, OperationType.OVERWRITE, or
        OperationType.DELETE)
        """
        # To be implemented by subclasses
        pass

    def create_label_form(self, operation_type):
        """
        This method should be overridden by subclasses to implement specific logic

        :param operation_type: The specific operation type (OperationType.Create, OperationType.OVERWRITE, or
        OperationType.DELETE)
        """
        # To be implemented by subclasses
        pass

    def submit_file_data(self, target_dict, operation_type):
        """
        Validates data collected from the File Information form and creates the Label Information form if operation_type
        isn't DELETE.

        :param target_dict: The dictionary containing all the data collected from the File Information form
        :type target_dict: dict
        :param operation_type: The specific operation type (OperationType.Create, OperationType.OVERWRITE, or
        OperationType.DELETE)
        :return: None
        """

        for label, var in self.file_widgets:
            if var.get().strip() == "":
                messagebox.showerror("Error", "There are missing fields in the current form")
                return
            target_dict[label] = var.get()

        # Validate the user-inputted Excel file
        if not validate_file_location(target_dict["Input File"]):
            messagebox.showerror("Error", "Invalid file input")
            return

        # Validate that the file is not open
        if is_file_open(target_dict["Input File"]):
            messagebox.showerror("Error", "Excel file is currently open. Please close it and try again")
            return

        # Validate that the sheet is within the user-inputted Excel file
        if sheet_exists(target_dict["Input File"], target_dict["Sheet Name"]):
            sheet_index = get_sheet_index(target_dict["Input File"], target_dict["Sheet Name"])

            # Replace 'Sheet Name' with 'Sheet Index' for simplicity
            target_dict["Sheet Index"] = target_dict.pop("Sheet Name")
            target_dict["Sheet Index"] = sheet_index
        else:
            messagebox.showerror("Error", "Invalid sheet name")
            return

        # Validate Column Letters
        if not is_valid_column(target_dict["Part Column Letter"]):
            messagebox.showerror("Error", "Invalid part column letter")
            return
        if "Description Column Letter" in target_dict:
            if target_dict["Description Column Letter"] != target_dict["Part Column Letter"]:
                if not is_valid_column(target_dict["Description Column Letter"]):
                    messagebox.showerror("Error", "Invalid description column letter")
                    return
            else:
                messagebox.showerror("Error", "Invalid description column letter")
                return

        # Validate row order
        if not is_valid_row_combo(target_dict["First Row"], target_dict["Last Row"]):
            messagebox.showerror("Error", "Invalid row or row combination. Max row count of 150")
            return

        # Validate that each column has no empty cells
        empty_rows = check_empty_rows(target_dict['Input File'], target_dict['Sheet Index'],
                                      target_dict['Part Column Letter'], int(target_dict['First Row']),
                                      int(target_dict['Last Row']))
        if len(empty_rows) > 0:
            messagebox.showerror("Error", f"There are empty cells in column "
                                          f"{target_dict['Part Column Letter']}. Please remove them and try again.")
            return

        if 'Description Column Letter' in target_dict:
            empty_rows.clear()
            empty_rows = check_empty_rows(target_dict['Input File'], target_dict['Sheet Index'],
                                          target_dict['Description Column Letter'], int(target_dict['First Row']),
                                          int(target_dict['Last Row']))
            if len(empty_rows) > 0:
                messagebox.showerror("Error", f"There are empty cells in column "
                                              f"{target_dict['Description Column Letter']}. "
                                              f"Please remove them and try again.")
                return

        # Verify the current subclass isn't DeleteForm
        class_name = type(self).__name__
        if class_name != "DeleteForm":
            self.create_label_form(operation_type)
        else:
            # Close the current form
            self.master.destroy()

    def submit_label_data(self, target_dict, operation_type):
        """
        Validates data collected from the Label Information form and terminates the self.master form

        :param target_dict: The dictionary containing all the data collected from the File Information form
        :type target_dict: dict
        :param operation_type: The specific operation type (OperationType.Create, OperationType.OVERWRITE, or
        OperationType.DELETE)
        :return: None
        """

        is_create_operation = False  # Variable that checks for the CREATE operation type
        if operation_type.name == "CREATE":
            is_create_operation = True

        empty_fields = 0  # Variable that counts the amount of total empty fields
        empty_dropdown_fields = 0  # Variable that counts the amount of empty dropdown fields

        for label, var in self.label_widgets:
            if var.get() == "" or var.get() is False:
                empty_fields += 1

            if var.get() == "":
                empty_dropdown_fields += 1

            target_dict[label] = var.get()

        # Check for empty dropdown fields in case of user creating
        if is_create_operation and empty_dropdown_fields > 0:
            messagebox.showerror("Input Error", "When you are creating parts you must fill in all dropdown"
                                                " fields")
            return

        # In the case of user overwriting with no inputs, ask for confirmation
        if empty_fields == len(target_dict):
            if not messagebox.askyesno("Warning", "You haven't made any changes. "
                                                  "Are you sure you want to proceed?"):
                return

        self.master.destroy()


class CreateForm(BaseForm):

    # Build custom File Form
    def create_file_form(self, operation_type):
        """
        Creates the File Information form using the CREATE operation type and links the form data to self.submit_file_data

        :param operation_type: The specific operation type (OperationType.Create, OperationType.OVERWRITE, or
        OperationType.DELETE)
        :return: None
        """

        self.master.title("File Information - Create")
        self.master.minsize(420, 240)
        self.first_frame = ttk.Frame(self.master, padding="10")
        self.first_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.create_file_widget(self.first_frame, "Input File", 0, 0, self.file_widgets)
        self.create_entry_widget(self.first_frame, "Sheet Name", 1, 0, self.file_widgets)
        self.create_entry_widget(self.first_frame, "Part Column Letter", 2, 0, self.file_widgets)
        self.create_entry_widget(self.first_frame, "Description Column Letter", 3, 0, self.file_widgets)
        self.create_entry_widget(self.first_frame, "First Row", 4, 0, self.file_widgets)
        self.create_entry_widget(self.first_frame, "Last Row", 5, 0, self.file_widgets)
        tk.Button(self.first_frame, text="Submit",
                  command=lambda: self.submit_file_data(self.file_data, operation_type)).grid(row=5,
                                                                                              column=2,
                                                                                              padx=(
                                                                                                  0,
                                                                                                  10),
                                                                                              pady=7)

    # Build custom Label Form
    def create_label_form(self, operation_type):
        """
        Creates the Label Information form using the CREATE operation type and links the form data to
        self.submit_label_data

        :param operation_type: The specific operation type (OperationType.Create, OperationType.OVERWRITE, or
        OperationType.DELETE)
        :return: None
        """

        self.master.title("Label Information - Create")
        self.second_frame = ttk.Frame(self.master, padding="10")
        self.second_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.create_dropdown_widget(self.second_frame, "Type", 13, TYPE_OPTIONS,
                                    0, 0, self.label_widgets)
        self.create_dropdown_widget(self.second_frame, "On Hold Reason", 28, ON_HOLD_REASON_OPTIONS,
                                    5, 0, self.label_widgets)
        self.create_dropdown_widget(self.second_frame, "Group", 28, GROUP_OPTIONS,
                                    1, 0, self.label_widgets)
        self.create_dropdown_widget(self.second_frame, "Class", 28, CLASS_OPTIONS,
                                    2, 0, self.label_widgets)
        self.create_dropdown_widget(self.second_frame, "Label Group", 28, LABEL_GROUP_OPTIONS,
                                    3, 0, self.label_widgets)
        self.create_dropdown_widget(self.second_frame, "Reporting Group", 28, REPORTING_GROUP_OPTIONS,
                                    4, 0, self.label_widgets)

        self.create_checkbox_widget(self.second_frame, "Priced Part", 0, 2, self.label_widgets)
        self.create_checkbox_widget(self.second_frame, "Salesforce Sync", 1, 2, self.label_widgets)
        self.create_checkbox_widget(self.second_frame, "Catalog Part", 2, 2, self.label_widgets)

        tk.Button(self.second_frame, text="Submit",
                  command=lambda: self.submit_label_data(self.label_data, operation_type)).grid(row=5, column=2, padx=(0
                                                                                                                       ,
                                                                                                                       10),
                                                                                                pady=7, )


class OverwriteForm(BaseForm):

    # Build custom File Form
    def create_file_form(self, operation_type):
        """
        Creates the File Information form using the OVERWRITE operation type and links the form data to self.submit_file_data

        :param operation_type: The specific operation type (OperationType.Create, OperationType.OVERWRITE, or
        OperationType.DELETE)
        :return: None
        """

        self.master.title("File Information - Overwrite")
        self.master.minsize(410, 200)
        self.first_frame = ttk.Frame(self.master, padding="10")
        self.first_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.create_file_widget(self.first_frame, "Input File", 0, 0, self.file_widgets)
        self.create_entry_widget(self.first_frame, "Sheet Name", 1, 0, self.file_widgets)
        self.create_entry_widget(self.first_frame, "Part Column Letter", 2, 0, self.file_widgets)
        self.create_entry_widget(self.first_frame, "First Row", 3, 0, self.file_widgets)
        self.create_entry_widget(self.first_frame, "Last Row", 4, 0, self.file_widgets)
        tk.Button(self.first_frame, text="Submit",
                  command=lambda: self.submit_file_data(self.file_data, operation_type)).grid(row=4,
                                                                                              column=2,
                                                                                              padx=(
                                                                                                  0,
                                                                                                  10)
                                                                                              , pady=7)

    # Build custom Label Form
    def create_label_form(self, operation_type):
        """
        Creates the Label Information form using the OVERWRITE operation type and links the form data to
        self.submit_label_data

        :param operation_type: The specific operation type (OperationType.Create, OperationType.OVERWRITE, or
        OperationType.DELETE)
        :return: None
        """

        self.master.title("Label Information - Overwrite")
        self.second_frame = ttk.Frame(self.master, padding="10")
        self.second_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.create_dropdown_widget(self.second_frame, "Type", 13, TYPE_OPTIONS,
                                    0, 0, self.label_widgets)
        self.create_dropdown_widget(self.second_frame, "On Hold Reason", 28, ON_HOLD_REASON_OPTIONS,
                                    5, 0, self.label_widgets)
        self.create_dropdown_widget(self.second_frame, "Group", 28, GROUP_OPTIONS,
                                    1, 0, self.label_widgets)
        self.create_dropdown_widget(self.second_frame, "Class", 28, CLASS_OPTIONS,
                                    2, 0, self.label_widgets)
        self.create_dropdown_widget(self.second_frame, "Label Group", 28, LABEL_GROUP_OPTIONS,
                                    3, 0, self.label_widgets)
        self.create_dropdown_widget(self.second_frame, "Reporting Group", 28, REPORTING_GROUP_OPTIONS,
                                    4, 0, self.label_widgets)

        self.create_checkbox_widget(self.second_frame, "Priced Part", 0, 2, self.label_widgets)
        self.create_checkbox_widget(self.second_frame, "Salesforce Sync", 1, 2, self.label_widgets)
        self.create_checkbox_widget(self.second_frame, "Catalog Part", 2, 2, self.label_widgets)

        tk.Button(self.second_frame, text="Submit", command=lambda: self.submit_label_data(self.label_data,
                                                                                           operation_type)).grid(row=5,
                                                                                                                 column=2,
                                                                                                                 padx=(
                                                                                                                     0,
                                                                                                                     10),
                                                                                                                 pady=7)


class DeleteForm(BaseForm):

    # Build custom File Form
    def create_file_form(self, operation_type):
        """
        Creates the File Information form using the DELETE operation type and links the form data to self.submit_file_data

        :param operation_type: The type of specified operation (CREATE, OVERWRITE, or DELETE)
        :return: None
        """

        self.master.title("File Information - Delete")
        self.master.minsize(410, 200)
        self.first_frame = ttk.Frame(self.master, padding="10")
        self.first_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.create_file_widget(self.first_frame, "Input File", 0, 0, self.file_widgets)
        self.create_entry_widget(self.first_frame, "Sheet Name", 1, 0, self.file_widgets)
        self.create_entry_widget(self.first_frame, "Part Column Letter", 2, 0, self.file_widgets)
        self.create_entry_widget(self.first_frame, "First Row", 3, 0, self.file_widgets)
        self.create_entry_widget(self.first_frame, "Last Row", 4, 0, self.file_widgets)
        tk.Button(self.first_frame, text="Submit", command=lambda: self.submit_file_data(self.file_data, operation_type)
                  ).grid(row=4, column=2, padx=(0, 10), pady=7)
