from enum import Enum
from abc import ABC, abstractmethod
import pywinauto.findwindows
from pywinauto import Application
from pywinauto.keyboard import send_keys
from pywinauto.timings import Timings
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import datetime
from datetime import datetime
import sys
import os
import shutil


def print_fancy_separator(text="", char='-'):
    """
    This function generates a visually appealing separator in the terminal.

    :param text: The text to be centered within the separator. Default is an empty string.
    :param char: The character used to build the separator. Default is '-'.
    :return: None
    """

    terminal_width, _ = shutil.get_terminal_size()
    if text:
        text = f" {text} "
    separator_width = (terminal_width - len(text)) // 2
    print(f"{char * separator_width}{text}{char * separator_width}")


class OperationType(Enum):
    # Shared dictionaries
    file_data = {}
    label_data = {}

    CREATE = 1
    OVERWRITE = 2
    DELETE = 3


class OperationLogger:
    def __init__(self):
        """
            Initializes the OperationLogger class instance.

            If the operations log file does not exist, a new workbook is created with predefined settings:
            - Filename is generated with the current date and time.
            - Workbook, active sheet, and sheet title are initialized.
            - Column headers for 'Operation', 'Part Number', 'Description', 'Status', and 'Timestamp' are defined.
            - The first row is made bold.
            - Column widths set based on specified lengths.

            If the operations log file already exists, it loads the existing workbook and sets the active sheet.
            """

        self.filename = f"operations_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        if not os.path.exists(self.filename):
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "Operations Log"

            # Define column headers
            headers = ["Operation", "Part Number", "Description", "Status", "Timestamp"]
            self.sheet.append(headers)

            # Make the first row bold
            for cell in self.sheet[1]:
                cell.font = Font(bold=True)

            # Set column widths using the specified lengths
            column_widths = [10, 12, 11, 11, 20]  # Using your specified lengths
            for i, column_width in enumerate(column_widths, 1):
                self.sheet.column_dimensions[get_column_letter(i)].width = column_width

            self.save_workbook()
        else:
            self.workbook = load_workbook(self.filename)
            self.sheet = self.workbook.active

    def log_operation(self, operation, part_number, description, status):
        """
        Write in the operation, part number, description, and status into a pre-made Excel spreadsheet

        :param operation: The specific operation being performed
        :type operation: str
        :param part_number: The specific part number being logged
        :type part_number: str
        :param description: The specific description being logged. If overwriting/deleting description is n/a.
        :param status: The status of the operation (Complete/Incomplete) and why

        :return: None
        """
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.sheet.append([operation, part_number, description, status, timestamp])
        self.save_workbook()

    def save_workbook(self):
        """
        Saves workbook
        :return: None
        """
        self.workbook.save(self.filename)


class Operation(ABC):
    @abstractmethod
    def execute(self, file_data, label_data):
        """
        This method is an abstract method that must be implemented by subclasses.

        :param file_data: A dictionary containing user data related to the file information form
        :type file_data: dict
        :param label_data: A dictionary containing user data related to the label information form
        :type label_data: dict

        This method serves as a template for executing specific operations. Subclasses of the 'Operation' class
        must implement this method to define the behavior of the operation they represent. It takes in 'file_data'
        and 'label_data' as parameters, which provide the necessary information for carrying out the operation.
        """
        pass


# Create a global instance of OperationLogger
operation_logger = OperationLogger()


class CreateOperation(Operation):
    def execute(self, file_data, label_data):
        """
        Execute method specific to the CreateOperation subclass.

        :param file_data: A dictionary containing user data related to the file information form
        :type file_data: dict
        :param label_data: A dictionary containing user data related to the label information form
        :type label_data: dict

        This method is implemented in the CreateOperation subclass to carry out the create operation. It takes in 'file_data'
        and 'label_data' as parameters, providing the necessary information for creating parts in the system.
        The method first establishes a connection to the Part Maintenance application, clears current information,
        loads the user-provided workbook, and loops through the part numbers to overwrite them. If a part number is null,
        it logs an incomplete operation. If the part doesn't exist, it logs that the creation cannot be performed.
        If the part exists, it proceeds with the creating process by confirming in Epicor.
        Upon successful overwriting, it logs the completion of the operation.
        """

        # Print messages and separators to the console
        print_fancy_separator("User Data")
        print(f"File Data: {file_data}\nLabel Data: {label_data}")
        print_fancy_separator("Program Documentation")
        print("Initializing Part Deletion...\n")

        try:
            # Connect the application to Part Maintenance and send confirmation message
            app = Application(backend="uia").connect(title="Part Maintenance")
            print('Connection to Part Maintenance achieved!\n')

            # Clear current information
            app.window(title='Part Maintenance').child_window(title="Clear").click_input()

            # Load in the user-provided workbook and access the specific sheet
            book = load_workbook(file_data["Input File"])
            sheet_index = file_data["Sheet Index"]
            sheets = book.sheetnames
            active_sheet = book[sheets[sheet_index]]

            # Loop through all the part numbers
            for i in range(int(file_data["First Row"]), int(file_data["Last Row"]) + 1):
                # Reconnect to the form toe ensure it doesn't fall asleep
                app = Application(backend="uia").connect(title="Part Maintenance")
                main_window = app.window(title='Part Maintenance')

                # Find the cell that hosts the part numbers and descriptions
                pn_cell = file_data["Part Column Letter"] + str(i)
                part_number = active_sheet[pn_cell].value
                desc_cell = file_data["Description Column Letter"] + str(i)
                part_description = active_sheet[desc_cell].value

                # Type cell value into text box
                main_window.child_window(auto_id='tbPart').type_keys(part_number)
                send_keys("{TAB}")

                # Validate that part number is not None
                if part_number is None:
                    operation_logger.log_operation("Create", str(part_number), part_description,
                                                   "Incomplete: part number was null")
                    print(str(part_number) + " - Unable to create: Part number is null")
                    main_window.child_window(auto_id='btnNo2').click_input()
                    continue

                # Confirm that the part does not already exist
                if main_window.child_window(title="Add New Confirmation").exists():
                    main_window.child_window(auto_id='btnYes2').click_input()
                    # Validate that part description is not None
                    if part_description is None:
                        operation_logger.log_operation("Create", str(part_number), part_description,
                                                       "Completed with empty description")
                        print(str(part_number) + " - Part Created   **No Description**")
                    else:
                        operation_logger.log_operation("Create", str(part_number), part_description,
                                                       "Completed")
                        print(str(part_number) + " - Part Created")
                else:
                    # Write PN into Excel file
                    operation_logger.log_operation("Create", str(part_number), part_description,
                                                   "Incomplete - Part already exists")
                    print(str(part_number) + " - Unable to create: Part already exists")
                    continue

                # Begin writing data into Epicor
                main_window.child_window(auto_id="tbPartDescription").type_keys(part_description, with_spaces=True)
                main_window.child_window(auto_id="cboTypeCode").type_keys(label_data["Type"], with_spaces=True)
                main_window.child_window(auto_id="cbProdCode").type_keys(label_data["Group"], with_spaces=True)
                main_window.child_window(auto_id="cbClass").type_keys(label_data["Class"], with_spaces=True)
                main_window.child_window(auto_id="ucbLabelGroup").type_keys(label_data["Label Group"],
                                                                            with_spaces=True)
                main_window.child_window(auto_id="cboReportGroup").type_keys(label_data["Reporting Group"],
                                                                             with_spaces=True)
                main_window.child_window(auto_id="cbOnHoldReasonCode").type_keys(label_data["On Hold Reason"],
                                                                                 with_spaces=True)

                # Here, we use simple logic to determine whether a checkbox should be clicked
                # Either the box is checked in our form and unchecked in Epicor or it's unchecked in our form and
                # checked in Epicor
                if (label_data["Priced Part"] and main_window.child_window(auto_id="epiCheckBox1").
                        get_toggle_state() == 0):
                    main_window.child_window(auto_id="epiCheckBox1").click_input()
                elif (not label_data["Priced Part"] and main_window.child_window(auto_id="epiCheckBox1").
                        get_toggle_state() == 1):
                    main_window.child_window(auto_id="epiCheckBox1").click_input()

                if (label_data["Salesforce Sync"] and main_window.child_window(auto_id="epiCheckBox2").
                        get_toggle_state() == 0):
                    main_window.child_window(auto_id="epiCheckBox2").click_input()
                elif (not label_data["Salesforce Sync"] and main_window.child_window(auto_id="epiCheckBox2").
                        get_toggle_state() == 1):
                    main_window.child_window(auto_id="epiCheckBox2").click_input()

                if (label_data["Catalog Part"] and main_window.child_window(auto_id="chkCatalogPart").
                        get_toggle_state() == 0):
                    main_window.child_window(auto_id="chkCatalogPart").click_input()
                elif (not label_data["Catalog Part"] and main_window.child_window(auto_id="chkCatalogPart").
                        get_toggle_state() == 1):
                    main_window.child_window(auto_id="chkCatalogPart").click_input()

                # Save the form and check for any unexpected errors
                main_window.child_window(title="Save").click_input()
                if main_window.child_window(title="Error").exists():
                    messagebox.showerror(
                        "Error",
                        "If you are creating parts and not overwriting existing ones, you must add a "
                        "description in the first form of the program. "
                    )
                main_window.child_window(title="Clear").click_input()

        except pywinauto.findwindows.ElementNotFoundError:
            print("Epicor Connection Failed...")
            messagebox.showinfo("Connection Failed", "Part Maintenance not found. \nTerminating "
                                                     "program...")
            sys.exit()
        except pywinauto.timings.TimeoutError:
            messagebox.showerror("Error", "The program took too long to respond. Please restart")
        except Exception as e:
            print(e)
            raise e


class OverwriteOperation(Operation):
    def execute(self, file_data, label_data):
        """
        Execute method specific to the OverwriteOperation subclass.

        :param file_data: A dictionary containing user data related to the file information form
        :type file_data: dict
        :param label_data: A dictionary containing user data related to the label information form
        :type label_data: dict

        This method is implemented in the OverwriteOperation subclass to carry out the overwriting operation. It takes in 'file_data'
        and 'label_data' as parameters, providing the necessary information for overwriting parts in the system.
        The method first establishes a connection to the Part Maintenance application, clears current information,
        loads the user-provided workbook, and loops through the part numbers to overwrite them. If a part number is null,
        it logs an incomplete operation. If the part doesn't exist, it logs that the overwriting cannot be performed.
        If the part exists, it proceeds with the overwriting process by confirming the edit in Epicor.
        Upon successful overwriting, it logs the completion of the operation.
        """

        # Print messages and separators to the console
        print_fancy_separator("User Data")
        print(f"File Data: {file_data}\nLabel Data: {label_data}")
        print_fancy_separator("Program Documentation")
        print("Initializing Part Deletion...\n")

        try:
            # Connect the application to Part Maintenance and send confirmation message
            app = Application(backend="uia").connect(title="Part Maintenance")
            print('Connection to Part Maintenance achieved\n')

            # Clear current information
            app.window(title='Part Maintenance').child_window(title="Clear").click_input()

            # Load in the user-provided workbook and access the specific sheet
            book = load_workbook(file_data["Input File"])
            sheet_index = file_data["Sheet Index"]
            sheets = book.sheetnames
            active_sheet = book[sheets[sheet_index]]

            # Loop through all the part numbers
            for i in range(int(file_data["First Row"]), int(file_data["Last Row"]) + 1):
                # Reconnect to the form toe ensure it doesn't fall asleep
                app = Application(backend="uia").connect(title="Part Maintenance")
                main_window = app.window(title='Part Maintenance')

                # Find the cell that hosts the part numbers
                pn_cell = file_data["Part Column Letter"] + str(i)
                part_number = active_sheet[pn_cell].value

                # Type cell value into text box
                main_window.child_window(auto_id='tbPart').type_keys(part_number)
                send_keys("{TAB}")

                # Validate that part number is not None
                if part_number is None:
                    operation_logger.log_operation("Create", str(part_number), "n/a",
                                                   "Incomplete: part number was null")
                    print(str(part_number) + " - Unable to overwrite: Part number is null")
                    main_window.child_window(auto_id='btnNo2').click_input()
                    continue

                # Confirm that the part already exist
                if main_window.child_window(title="Add New Confirmation").exists():
                    main_window.child_window(auto_id='btnNo2').click_input()
                    operation_logger.log_operation("Overwrite", part_number, "n/a", "Incomplete - "
                                                                                    "part doesn't exist and therefore "
                                                                                    "can't"
                                                                                    "be overwritten")
                    print(str(part_number) + " - Unable to overwrite: Part never existed")
                    continue

                # Conditionally write in any existing fields into Epicor
                if label_data["Type"]:
                    main_window.child_window(auto_id="cboTypeCode").type_keys(label_data["Type"], with_spaces=True)
                if label_data["Group"]:
                    main_window.child_window(auto_id="cbProdCode").type_keys(label_data["Group"], with_spaces=True)
                if label_data["Class"]:
                    main_window.child_window(auto_id="cbClass").type_keys(label_data["Class"], with_spaces=True)
                if label_data["Label Group"]:
                    main_window.child_window(auto_id="ucbLabelGroup").type_keys(label_data["Label Group"],
                                                                                with_spaces=True)
                if label_data["Reporting Group"]:
                    main_window.child_window(auto_id="cboReportGroup").type_keys(label_data["Reporting Group"],
                                                                                 with_spaces=True)
                if label_data["On Hold Reason"]:
                    main_window.child_window(auto_id="cbOnHoldReasonCode").type_keys(label_data["On Hold Reason"],
                                                                                     with_spaces=True)

                # Here, we use simple logic to determine whether a checkbox should be clicked
                # Either the box is checked in our form and unchecked in Epicor or it's unchecked in our form and
                # checked in Epicor
                if label_data["Priced Part"] and main_window.child_window(auto_id="epiCheckBox1").get_toggle_state() \
                        == 0:
                    main_window.child_window(auto_id="epiCheckBox1").click_input()
                elif (not label_data["Priced Part"] and main_window.child_window(auto_id="epiCheckBox1").
                        get_toggle_state() == 1):
                    main_window.child_window(auto_id="epiCheckBox1").click_input()

                if (label_data["Salesforce Sync"] and main_window.child_window(auto_id="epiCheckBox2").
                        get_toggle_state() == 0):
                    main_window.child_window(auto_id="epiCheckBox2").click_input()
                elif (not label_data["Salesforce Sync"] and main_window.child_window(auto_id="epiCheckBox2").
                        get_toggle_state() == 1):
                    main_window.child_window(auto_id="epiCheckBox2").click_input()

                if (label_data["Catalog Part"] and main_window.child_window(auto_id="chkCatalogPart").
                        get_toggle_state() == 0):
                    main_window.child_window(auto_id="chkCatalogPart").click_input()
                elif (not label_data["Catalog Part"] and main_window.child_window(auto_id="chkCatalogPart").
                        get_toggle_state() == 1):
                    main_window.child_window(auto_id="chkCatalogPart").click_input()

                # Save the form and check for any unexpected errors
                main_window.child_window(title="Save").click_input()
                if main_window.child_window(title="Error").exists():
                    messagebox.showerror(
                        "Error",
                        "An error has occurred. Please try again."
                    )

                # Confirm saving
                if main_window.child_window(title="Save Confirmation").exists():
                    confirmation_dialog = main_window.child_window(title="Save Confirmation",
                                                                   auto_id="EpiCheckMessageBox")
                    yes_button = confirmation_dialog.child_window(title="Yes", auto_id="btnYes2", control_type="Button")
                    yes_button.click_input()

                # Log successful operation
                operation_logger.log_operation("Overwrite", part_number, "n/a", "Completed")
                print(str(part_number) + " - Overwrite Complete")

                # Clear form
                main_window.child_window(title="Clear").click_input()

        except pywinauto.findwindows.ElementNotFoundError as e:
            print("Epicor Connection Failed...")
            messagebox.showinfo("Connection Failed", "Part Maintenance not found. \nTerminating "
                                                     "program...")
            sys.exit()
        except pywinauto.timings.TimeoutError:
            messagebox.showerror("Error", "The program took too long to respond. Please restart")
        except Exception as e:
            print(e)
            raise e


class DeleteOperation(Operation):
    def execute(self, file_data, label_data):
        """
        Execute method specific to the DeleteOperation subclass.

        :param file_data: A dictionary containing user data related to the file information form
        :type file_data: dict
        :param label_data: A dictionary containing user data related to the label information form
        :type label_data: dict

        This method is implemented in the DeleteOperation subclass to carry out the deletion operation. It takes in 'file_data'
        and 'label_data' as parameters, providing the necessary information for deleting parts in the system.
        The method first establishes a connection to the Part Maintenance application, clears current information,
        loads the user-provided workbook, and loops through the part numbers to delete them. If a part number is null,
        it logs an incomplete operation. If the part doesn't exist, it logs that the deletion cannot be performed.
        If the part exists, it proceeds with the deletion process by confirming the deletion in Epicor.
        Upon successful deletion, it logs the completion of the operation.
        """

        # Print messages and separators to the console
        print_fancy_separator("User Data")
        print(f"File Data: {file_data}\nLabel Data: {label_data}")
        print_fancy_separator("Program Documentation")
        print("Initializing Part Deletion...\n")

        try:
            # Connect the application to Part Maintenance and send confirmation message
            app = Application(backend="uia").connect(title="Part Maintenance")
            print('Connection to Part Maintenance achieved\n')

            # Clear current information
            app.window(title='Part Maintenance').child_window(title="Clear").click_input()

            # Load in the user-provided workbook and access the specific sheet
            book = load_workbook(file_data["Input File"])
            sheet_index = file_data["Sheet Index"]
            sheets = book.sheetnames
            active_sheet = book[sheets[sheet_index]]

            # Loop through all the part numbers
            for i in range(int(file_data["First Row"]), int(file_data["Last Row"]) + 1):
                # Reconnect to the form toe ensure it doesn't fall asleep
                app = Application(backend="uia").connect(title="Part Maintenance")
                main_window = app.window(title='Part Maintenance')

                # Find the cell that hosts the part numbers
                pn_cell = file_data["Part Column Letter"] + str(i)
                part_number = active_sheet[pn_cell].value

                # Type cell value into text box
                main_window.child_window(auto_id='tbPart').type_keys(part_number)
                send_keys("{TAB}")

                # Validate that part number is not None
                if part_number is None:
                    operation_logger.log_operation("Create", str(part_number), "n/a",
                                                   "Incomplete: part number was null")
                    print(str(part_number) + " - Unable to delete: Part number is null")
                    main_window.child_window(auto_id='btnNo2').click_input()
                    continue

                # Confirm that the part already exist
                if main_window.child_window(title="Add New Confirmation").exists():
                    main_window.child_window(auto_id='btnNo2').click_input()
                    operation_logger.log_operation("Delete", part_number, "n/a", "Incomplete - "
                                                                                 "part doesn't exist and therefore "
                                                                                 "can't be deleted")
                    print(str(part_number) + " - Unable to delete: Part never existed")
                    continue
                else:
                    main_window.child_window(title="Delete").click_input()
                    if main_window.child_window(title="Delete Confirmation").exists():
                        main_window.child_window(auto_id='btnYes2').click_input()
                        operation_logger.log_operation("Delete", part_number, "n/a", "Completed")
                        print(str(part_number) + " - Deletion Complete")

        except pywinauto.findwindows.ElementNotFoundError:
            print("Epicor Connection Failed...")
            messagebox.showinfo("Connection Failed", "Part Maintenance not found. \nTerminating "
                                                     "program...")
            sys.exit()
        except pywinauto.timings.TimeoutError:
            messagebox.showerror("Error", "The program took too long to respond. Please restart")
        except Exception as e:
            print(e)
            raise e


class ERPManager:
    def __init__(self, create_op: Operation, overwrite_op: Operation, delete_op: Operation):
        """
        Initializes an instance of the ERPManager class with operations for create, overwrite, and delete

        :param create_op: Operation object for the CREATE operation
        :param overwrite_op: Operation object for the OVERWRITE operation
        :param delete_op: Operation object for the DELETE operation

        This method sets up a dictionary 'operations' where keys are OperationType enums
        (CREATE, OVERWRITE, DELETE) and values are the corresponding operation objects.
        """

        self.operations = {
            OperationType.CREATE: create_op,
            OperationType.OVERWRITE: overwrite_op,
            OperationType.DELETE: delete_op
        }

    def perform_operation(self, op_type: OperationType, form_data, label_data):
        """
        Perform the specified operation based on the given operation type.

        :param op_type: The type of operation to be performed (CREATE, OVERWRITE, DELETE)
        :type op_type: OperationType
        :param form_data: Data related to the form for the operation
        :param label_data: Data related to the labels for the operation

        :raises ValueError: If the provided operation type is not valid

        This method retrieves the operation based on the operation type from the 'operations' dictionary
        in the ERPManager instance and executes the operation with the provided form_data and label_data.
        If the operation type is not found in the dictionary, a ValueError is raised.
        """
        operation = self.operations.get(op_type)
        if operation:
            operation.execute(form_data, label_data)
        else:
            raise ValueError("Invalid operation type")
