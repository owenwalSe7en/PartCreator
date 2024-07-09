import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, exceptions
from openpyxl.styles import Font
import openpyxl
from datetime import datetime
import os
import re
import sys
from pywinauto import Application
from pywinauto.keyboard import send_keys

wb = Workbook()
ws = wb.active
# File Data Layout: [input file, sheet index, PN column letter, description column letter, first row, last row]
file_data = {}

# Label Data Layout: [Type, Group, Class, Label Group, Reporting Group, On Hold Reason,
# Priced Part (Y/N), Salesforce Sync (Y/N), Catalog Part (Y/N), Kit Catalog (Y/N)]
label_data = {}


class Automator:
    def __init__(self):
        self.existed_index = 2

        # region UI Methods
        def browse_file():
            file_path = filedialog.askopenfilename()
            txtInputFile.delete(0, tk.END)
            txtInputFile.insert(0, file_path)

        def get_textbox_contents():
            if is_file_open(txtInputFile.get().strip()):
                messagebox.showerror("Error", "The file is currently open. Please close it and try again")
            else:
                try:
                    # Validate Input File
                    if txtInputFile.get():
                        if validate_file_location(txtInputFile.get().strip()):
                            file_data["Input File"] = txtInputFile.get().strip()
                        else:
                            messagebox.showerror("Error", "Invalid input file")
                            return

                        # Validate Sheet
                        if sheet_exists(txtInputFile.get(), txtSheetName.get().strip()):
                            index = get_sheet_index(txtInputFile.get().strip(), txtSheetName.get().strip())
                            file_data["Sheet Index"] = index
                        else:
                            messagebox.showerror("Error", "Invalid sheet name")
                            return
                    else:
                        messagebox.showerror("Error", "Please list an input file")
                        return

                    # Validate  Part Column Letter
                    if is_valid_column(txtColumnLetter_P.get().strip()):
                        file_data["PN Column Letter"] = txtColumnLetter_P.get().strip()

                        # Because Description column letter is optional we don't check for its input from the user
                        file_data["Desc Column Letter"] = txtColumnLetter_D.get().strip()
                    else:
                        messagebox.showerror("Error", "Invalid part column letter")
                        return

                    # Validate Rows
                    if is_valid_integer(txtFirstRow.get().strip()):
                        if is_valid_integer(txtLastRow.get().strip()):
                            if int(txtFirstRow.get()) < int(txtLastRow.get().strip()):
                                file_data["First Row"] = txtFirstRow.get().strip()
                            else:
                                messagebox.showerror("Error", "Invalid row order")
                                return
                        else:
                            messagebox.showerror("Error", "Last row is invalid")
                            return
                    else:
                        messagebox.showerror("Error", "First row is invalid")
                        return

                    file_data["Last Row"] = txtLastRow.get().strip()

                    print("File Data: " + str(file_data))
                    root.destroy()

                except Exception as e:
                    print(e)

        def is_valid_column(column):
            return re.match(r'^[A-Za-z]+$', column)

        def is_valid_integer(value):
            return value.isdigit()

        def sheet_exists(excel_file_path, sheet_name):
            try:
                workbook = openpyxl.load_workbook(excel_file_path)
                return sheet_name in workbook.sheetnames
            except FileNotFoundError:
                print(f"File not found: {excel_file_path}")
                return False
            except openpyxl.utils.exceptions.InvalidFileException:
                print(f"Invalid Excel file: {excel_file_path}")
                return False
            except Exception(BaseException):
                print("An error occured. Please try again.")

        def get_sheet_index(excel_file_path, sheet_name):
            try:
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_index = workbook.sheetnames.index(sheet_name)
                return sheet_index
            except FileNotFoundError:
                print(f"File not found: {excel_file_path}")
                return False
            except openpyxl.utils.exceptions.InvalidFileException:
                print(f"Invalid Excel file: {excel_file_path}")
                return False
            except ValueError:
                print(f"Sheet '{sheet_name}' not found in the Excel file.")
                return False

        def validate_file_location(file_path):
            # Check if the file path is not empty
            if not file_path:
                return False

            # Check if the file path exists
            if not os.path.exists(file_path):
                return False

            # Check if the file path points to a file (not a directory)
            if os.path.isdir(file_path):
                return False

            return True

        def is_file_open(filepath):
            try:
                os.rename(filepath, filepath)  # Attempt to rename the file to itself
                return False  # File is not open
            except OSError as e:
                return True  # File is open

        def on_closing():
            if messagebox.askokcancel("Quit", "Do you want to quit?"):
                root.destroy()
                sys.exit()

        def delete_all():
            try:
                get_part_information()

                # Connect to Epicor to ensure it is running
                app = Application(backend="uia").connect(title="Part Maintenance")
                print('Connection to Part Maintenance achieved')

                # Clear current information
                app.window(title='Part Maintenance').child_window(title="Clear").click_input()

                # Load in a specific workbook
                book = load_workbook(file_data["Input File"])

                # Load in user data
                last_row = int(file_data["Last Row"])
                first_row = int(file_data["First Row"])
                pn_col_letter = file_data["PN Column Letter"]
                n = file_data["Sheet Index"]
                sheets = book.sheetnames
                return_sheet = book[sheets[n]]

                for i in range(first_row, last_row + 1):

                    # Reconnect to the form to ensure it doesn't fall asleep
                    app = Application(backend="uia").connect(title="Part Maintenance")
                    main_window = app.window(title='Part Maintenance')

                    # Find the cells for the part number
                    pn_cell = pn_col_letter + str(i)
                    part_number = return_sheet[pn_cell].value

                    # Type cell value into text box
                    main_window.child_window(auto_id='tbPart').type_keys(part_number)
                    send_keys("{TAB}")

                    # Confirm that the part already exist
                    if main_window.child_window(title="Add New Confirmation").exists():
                        main_window.child_window(auto_id='btnNo2').click_input()
                        print(str(part_number) + " - part doesn't exist and therefore cannot be deleted")
                    else:
                        main_window.child_window(title="Delete").click_input()
                        if main_window.child_window(title="Delete Confirmation").exists():
                            main_window.child_window(auto_id='btnYes2').click_input()
                            print(str(part_number) + " - Part Deleted")
            except Exception as e:
                print(e)
                raise e
            sys.exit()

        # endregion UI Methods

        # region File Information Form
        root = tk.Tk()
        root.title("Label Creator - File Information")

        # Bind the window close event to the custom method
        root.protocol("WM_DELETE_WINDOW", on_closing)

        # Labels
        lblInputFile = tk.Label(root, text="Input File", padx=10, pady=10)
        lblSheetIndex = tk.Label(root, text="Sheet Name", padx=10, pady=10)
        lblColumnLetter_P = tk.Label(root, text="Column Letter (Part Number)", padx=10, pady=10)
        lblColumnLetter_D = tk.Label(root, text="Column Letter (Description)", padx=10, pady=10)
        lblFirstRow = tk.Label(root, text="First Row #", padx=10, pady=10)
        lblLastRow = tk.Label(root, text="Last Row #", padx=10, pady=10)

        # Note
        note_text1 = ("Note: If you are editing an existing part or deleting parts, you do not need to add the"
                      " description column letter")
        note_label1 = ttk.Label(root, text=note_text1, foreground="red", wraplength=580)

        # Text boxes
        txtInputFile = tk.Entry(root)
        txtSheetName = tk.Entry(root)
        txtColumnLetter_P = tk.Entry(root)
        txtColumnLetter_D = tk.Entry(root)
        txtFirstRow = tk.Entry(root)
        txtLastRow = tk.Entry(root)

        # Buttons
        browse_button = tk.Button(root, text="Browse", command=browse_file, padx=7, pady=7)
        submit_button = tk.Button(root, text="Submit", command=get_textbox_contents, padx=7, pady=7)

        # Grid layout
        lblInputFile.grid(row=0, column=0, sticky="w", padx=7, pady=7)
        txtInputFile.grid(row=0, column=1, padx=7, pady=7)
        lblSheetIndex.grid(row=1, column=0, sticky="w", padx=7, pady=7)
        txtSheetName.grid(row=1, column=1, padx=7, pady=7)
        lblColumnLetter_P.grid(row=2, column=0, sticky="w", padx=7, pady=7)
        txtColumnLetter_P.grid(row=2, column=1, padx=10, pady=10)
        lblColumnLetter_D.grid(row=3, column=0, sticky="w", padx=7, pady=7)
        txtColumnLetter_D.grid(row=3, column=1, padx=10, pady=10)
        lblFirstRow.grid(row=4, column=0, sticky="w", padx=7, pady=7)
        txtFirstRow.grid(row=4, column=1, padx=7, pady=7)
        lblLastRow.grid(row=5, column=0, sticky="w", padx=7, pady=7)
        txtLastRow.grid(row=5, column=1, padx=7, pady=7)
        browse_button.grid(row=0, column=2, padx=10, pady=7)
        submit_button.grid(row=5, column=2, padx=10, pady=7)
        note_label1.grid(row=6, column=0, columnspan=2, sticky=tk.W, pady=(10, 0))

        root.resizable(False, False)
        root.mainloop()

        # endregion UI Prompts

        # region Label Information Form

        root = tk.Tk()
        root.title("Label Creator - Label Information")

        # Bind the window close event to the custom method
        root.protocol("WM_DELETE_WINDOW", on_closing)

        # region ComboBox Options
        type_options = ["Manufactured", "Purchased", "Sales Kit"]
        on_hold_reason_options = [
            "Engineering - In-Process",
            "Engineering prototype",
            "Engineering Review Needed",
            "Marketing Updates",
            "Materials - FTT Review needed",
            "MOM Detail Review",
            "New Part - Needs review",
            "No Converted MOM",
            "Obsolete - Slow Moving Inv.",
            "Obsolete Part",
            "Part End of Life"
        ]
        group_options = [
            "_COMP - Component parts",
            "_CSUP - Cust. Supplied",
            "_CUST - Custom parts",
            "_DIST - Distributed parts",
            "_ENCL - Standard - Enclosure",
            "_EWARR Service - Ext. Warranty",
            "_FS-EPD -Service - Field - EPD",
            "_FS-MPD -Service - Field - MPD",
            "_HG7 - Technical - HG7 Panels",
            "_HG7-Kit - HG7 Kits",
            "_HGA-Technical-Active Filters",
            "_HGA - Technical - HGL Panels",
            "_HGL-KIT-Technical-HGL Kits",
            "_HGP - Technical - HGP Panels",
            "_HGP-KIT-Technical-HGP Kits",
            "_HSD-Technical-Harmonicshield",
            "_HSE",
            "_HSE-KIT - HSE Kits",
            "_HSL",
            "_KDR - Standard KDR",
            "_KDRL - KDRL",
            "_KDRTT - Standard KDRTT",
            "_KDRX - Standard KDRX",
            "_KLC - Standard - KLC",
            "_KLR - Standard - KLR",
            "_KMG - Technical - KMG",
            "_KRF",
            "_KTR - Standard - KTR",
            "_MFC - Technical - MFC",
            "_MSD - Technical - Motorshield",
            "_OBSDIST - Obsolete - DIST",
            "_OBSHF - Obsolete - HF",
            "_PASSIVE - Standard - Passive",
            "_PFG - Technical - PF Guard",
            "_R&D - R&D parts - Not Sold",
            "_V1K - Standard - V1K",
            "_V1K-KIT"
        ]
        class_options = [
            "_FG - Custom Products",
            "_FG - Distributed Products",
            "_FG - Electronic Products",
            "_FG - Harmonic Filter",
            "_FG - Output Filters",
            "_FG - Power Factor Correction",
            "_FG - Reactors",
            "_FG - Sinewave Filter",
            "Assembly",
            "Core",
            "Electrical",
            "Enclosures",
            "Extended Warranty - EPD",
            "Extended Warranty - MPD",
            "Fabricated Parts",
            "Field Service - EPD",
            "Hardware",
            "Insulators",
            "Labels",
            "Literature",
            "Miscellaneous",
            "Packaging",
            "R&D EPD parts",
            "Raw Metallic",
            "Shop Supplies",
            "Terminals",
            "Winding Conductors"
        ]
        label_group_optinos = [
            "EHFB",
            "EHFT",
            "EPF",
            "GE Oil and Gas Assembly",
            "HG",
            "HG7",
            "HarmonicGuard Low Capacitance Filter",
            "HarmonicGuard Passive Filter",
            "HarmonicShield",
            "HSE",
            "HSL",
            "KCAP",
            "KDR",
            "KDRUL",
            "KHG",
            "KLC",
            "KLCUL",
            "KLR",
            "KLRUL",
            "KMG",
            "KPC Capacitor Bank",
            "KTR",
            "KTRUL",
            "MFC",
            "MotorShield",
            "PCB",
            "PF Guard",
            "PUR or ENCL (box label only)",
            "V1K"
        ]
        reporting_group_options = [
            "_COMP - Component parts",
            "_CSUP - Cust. Supplied",
            "_CUST - Custom parts",
            "_DIST - Distributed parts",
            "_ENCL - Standard - Enclosure",
            "_EPF",
            "_EWARR Service - Ext. Warranty",
            "_FS-EPD -Service - Field - EPD",
            "_FS-MPD -Service - Field - MPD",
            "_HG7 - Technical - HG7 Panels",
            "_HG7-KIT - HG7 Kits",
            "_HGA-Technical-Active Filters",
            "_HGL - Technical - HGL Panels",
            "_HGL-KIT-Technical-HGL Kits",
            "_HGP - Technical - HGP Panels",
            "_HGP-KIT-Technical-HGP Kits",
            "_HSD-Technical-Harmonicshield",
            "_KDR - Standard KDR",
            "_KDRTT - Standard KDRTT",
            "_KLC - Standard - KLC",
            "_KLR - Standard - KLR",
            "_KMG - Technical - KMG",
            "_KTR - Standard - KTR",
            "_MFC - Technical - MFC",
            "_MSD - Technical - Motorshield",
            "_OBSDIST - Obsolete - DIST",
            "_OBSHF - Obsolete - HF",
            "_PASSIVE - Standard - Passive",
            "_PFG - Technical - PF Guard",
            "_R&D - R&D parts - Not Sold",
            "_V1K - Standard - V1K"
        ]
        # endregion

        # Labels
        lblGroup = tk.Label(root, text="Group", padx=10, pady=10)
        lblClass = tk.Label(root, text="Class", padx=10, pady=10)
        lblLabelGroup = tk.Label(root, text="Label Group", padx=10, pady=10)
        lblReportingGroup = tk.Label(root, text="Reporting Group", padx=10, pady=10)
        lblType = tk.Label(root, text="Type", padx=10, pady=10)
        lblOnHoldReason = tk.Label(root, text="On Hold Reason", padx=10, pady=10)

        # Note
        note_text2 = ("Note: If you are editing an existing part, you only need to change the intended control and the "
                      "checkboxes. The checkboxes always reflect what you will see in Epicor. If you are deleting "
                      "parts, just click the 'Delete All' Button.")
        note_label2 = ttk.Label(root, text=note_text2, foreground="red", wraplength=580)

        # Dropdown Menus/Combo boxes
        group_var = tk.StringVar()
        cboGroup = ttk.Combobox(root, textvariable=group_var, values=group_options, state="readonly")
        class_var = tk.StringVar()
        cboClass = ttk.Combobox(root, textvariable=class_var, values=class_options, state="readonly")
        lg_var = tk.StringVar()
        cboLabelGroup = ttk.Combobox(root, textvariable=lg_var, values=label_group_optinos, state="readonly")
        rg_var = tk.StringVar()
        cboReportingGroup = ttk.Combobox(root, textvariable=rg_var, values=reporting_group_options, state="readonly")
        type_var = tk.StringVar()
        cboType = ttk.Combobox(root, textvariable=type_var, values=type_options, state="readonly")
        ohr_var = tk.StringVar()
        cboOnHoldReason = ttk.Combobox(root, textvariable=ohr_var, values=on_hold_reason_options, state="readonly")

        # Check boxes
        PricedPart_value = tk.BooleanVar()
        chkPricedPart = tk.Checkbutton(root, text="Priced Part", variable=PricedPart_value)
        SalesforceSync_value = tk.BooleanVar()
        chkSalesforceSync = tk.Checkbutton(root, text="Salesforce Sync", variable=SalesforceSync_value)
        CatalogPart_value = tk.BooleanVar()
        chkCatalogPart = tk.Checkbutton(root, text="Catalog Part", variable=CatalogPart_value)
        KitCatalog_value = tk.BooleanVar()
        chkKitCatalog = tk.Checkbutton(root, text="Kit Catalog", variable=KitCatalog_value)

        # Placed here, so it can access the chk variables but also be called by the submit button
        def get_part_information():
            try:
                # Necessary inputs
                label_data["Type"] = cboType.get()
                label_data["Group"] = cboGroup.get()
                label_data["Class"] = cboClass.get()
                label_data["Label Group"] = cboLabelGroup.get()
                label_data["Reporting Group"] = cboReportingGroup.get()
                label_data["On Hold Reason"] = cboOnHoldReason.get()

                # Optional inputs
                label_data["Priced Part"] = PricedPart_value.get()
                label_data["Salesforce Sync"] = SalesforceSync_value.get()
                label_data["Catalog Part"] = CatalogPart_value.get()
                label_data["Kit Catalog"] = KitCatalog_value.get()

                print("Label Data: " + str(label_data))
                root.destroy()
            except Exception as e:
                print(e)
                raise e

        # Buttons
        submit_button = tk.Button(root, text="Submit", command=get_part_information, padx=7, pady=7)
        delete_button = tk.Button(root, text="Delete All", command=delete_all, padx=7, pady=7)

        # Grid Layout
        lblType.grid(row=0, column=0, sticky="w", padx=7, pady=6)
        lblOnHoldReason.grid(row=1, column=0, sticky="w", padx=7, pady=7)
        lblGroup.grid(row=2, column=0, sticky="w", padx=7, pady=6)
        lblClass.grid(row=3, column=0, sticky="w", padx=7, pady=6)
        lblLabelGroup.grid(row=4, column=0, sticky="w", padx=7, pady=6)
        lblReportingGroup.grid(row=5, column=0, sticky="w", padx=7, pady=6)
        cboType.grid(row=0, column=1, sticky="w", padx=10, pady=5)
        cboType.config(width=13)
        cboOnHoldReason.grid(row=1, column=1, sticky="w", padx=10, pady=5)
        cboGroup.grid(row=2, column=1, sticky="w", padx=10, pady=5)
        cboGroup.config(width=28)
        cboClass.grid(row=3, column=1, sticky="w", padx=10, pady=5)
        cboClass.config(width=28)
        cboLabelGroup.grid(row=4, column=1, sticky="w", padx=10, pady=5)
        cboLabelGroup.config(width=28)
        cboReportingGroup.grid(row=5, column=1, sticky="w", padx=10, pady=5)
        cboReportingGroup.config(width=28)
        cboOnHoldReason.config(width=28)
        chkPricedPart.grid(row=0, column=2, sticky="w", padx=7, pady=7)
        chkSalesforceSync.grid(row=1, column=2, sticky="w", padx=7, pady=7)
        chkCatalogPart.grid(row=2, column=2, sticky="w", padx=7, pady=7)
        chkKitCatalog.grid(row=3, column=2, sticky="w", padx=7, pady=7)
        submit_button.grid(row=5, column=2, padx=5, pady=7)
        delete_button.grid(row=4, column=2, padx=5, pady=7)
        note_label2.grid(row=6, column=0, columnspan=2, sticky=tk.W, pady=7)

        root.mainloop()
        # endregion

        self.file_name = Automator.create_excel(self)

    def automate(self):
        try:

            # Connect the application to Label Maintenance and send confirmation message
            app = Application(backend="uia").connect(title="Part Maintenance")
            print('Connection to Part Maintenance achieved')

            # Clear current information
            app.window(title='Part Maintenance').child_window(title="Clear").click_input()

            # Load in a specific workbook
            book = load_workbook(file_data["Input File"])

            # Load in user data
            last_row = int(file_data["Last Row"])
            first_row = int(file_data["First Row"])
            pn_col_letter = file_data["PN Column Letter"]\

            # Check for usage of the description column letter
            if file_data["Desc Column Letter"] != "":
                desc_col_letter = file_data["Desc Column Letter"]
            else:
                desc_col_letter = ''

            n = file_data["Sheet Index"]
            sheets = book.sheetnames
            return_sheet = book[sheets[n]]

            # Loop through all the part numbers
            for i in range(first_row, last_row + 1):

                # Reconnect to the form to ensure it doesn't fall asleep
                app = Application(backend="uia").connect(title="Part Maintenance")
                main_window = app.window(title='Part Maintenance')

                # Find the cells for the part number
                pn_cell = pn_col_letter + str(i)
                part_number = return_sheet[pn_cell].value

                # Check for usage of the description column letter
                if file_data["Desc Column Letter"] != "":
                    desc_cell = desc_col_letter + str(i)
                    part_desc = return_sheet[desc_cell].value
                else:
                    part_desc = ''

                # Type cell value into text box
                main_window.child_window(auto_id='tbPart').type_keys(part_number)
                send_keys("{TAB}")

                # Confirm that the part does not already exist
                if main_window.child_window(title="Add New Confirmation").exists():
                    main_window.child_window(auto_id='btnYes2').click_input()
                    print(str(part_number) + " - Part Created")
                else:
                    # Write PN into Excel file
                    ws['A' + str(self.existed_index)] = part_number
                    self.existed_index += 1
                    wb.save(self.file_name)
                    print(str(part_number) + " - Part Already Exists")

                # Write all the data into the Epicor form
                # Consider that each datapoint is optional and may be omitted if the intention is to overwrite
                if desc_col_letter != '':
                    main_window.child_window(auto_id="tbPartDescription").type_keys(part_desc, with_spaces=True)
                if label_data["Type"] != '':
                    main_window.child_window(auto_id="cboTypeCode").type_keys(label_data["Type"], with_spaces=True)
                if label_data["Group"] != '':
                    main_window.child_window(auto_id="cbProdCode").type_keys(label_data["Group"], with_spaces=True)
                if label_data["Class"] != '':
                    main_window.child_window(auto_id="cbClass").type_keys(label_data["Class"], with_spaces=True)
                if label_data["Label Group"] != '':
                    main_window.child_window(auto_id="ucbLabelGroup").type_keys(label_data["Label Group"],
                                                                                with_spaces=True)
                if label_data["Reporting Group"] != '':
                    main_window.child_window(auto_id="cboReportGroup").type_keys(label_data["Reporting Group"],
                                                                                 with_spaces=True)
                if label_data["On Hold Reason"] != '':
                    main_window.child_window(auto_id="cbOnHoldReasonCode").type_keys(label_data["On Hold Reason"],
                                                                                     with_spaces=True)

                # Here, we use simple logic to determine whether a checkbox should be clicked
                # Either the box is checked in our form and unchecked in Epicor or it's unchecked in our form and
                # checked in Epicor
                if label_data["Priced Part"] and main_window.child_window(auto_id="epiCheckBox1").get_toggle_state() \
                        == 0:
                    main_window.child_window(auto_id="epiCheckBox1").click_input()
                elif (not label_data["Priced Part"] and main_window.child_window(auto_id="epiCheckBox1").
                        get_toggle_state() == 1):
                    main_window.child_window(auto_id="epiCheckBox1").click_input()

                if (label_data["Salesforce Sync"] and main_window.child_window(auto_id="epiCheckBox2").
                        get_toggle_state() == 0):
                    main_window.child_window(auto_id="epiCheckBox2").click_input()
                elif (not label_data["Salesforce Sync"] and main_window.child_window(auto_id="epiCheckBox2").
                        get_toggle_state() == 1):
                    main_window.child_window(auto_id="epiCheckBox2").click_input()

                if (label_data["Catalog Part"] and main_window.child_window(auto_id="chkCatalogPart").
                        get_toggle_state() == 0):
                    main_window.child_window(auto_id="chkCatalogPart").click_input()
                elif (not label_data["Catalog Part"] and main_window.child_window(auto_id="chkCatalogPart").
                        get_toggle_state() == 1):
                    main_window.child_window(auto_id="chkCatalogPart").click_input()

                if (label_data["Kit Catalog"] and main_window.child_window(auto_id="chkKitCatalog").
                        get_toggle_state() == 0):
                    main_window.child_window(auto_id="chkKitCatalog").click_input()
                elif (not label_data["Kit Catalog"] and main_window.child_window(auto_id="chkKitCatalog").
                        get_toggle_state() == 1):
                    main_window.child_window(auto_id="chkKitCatalog").click_input()

                main_window.child_window(title="Save").click_input()
                if main_window.child_window(title="Error").exists():
                    messagebox.showerror(
                        "Error",
                        "If you are creating parts and not overwriting existing ones, you must add a "
                        "description in the first form of the program. "
                    )
                main_window.child_window(title="Clear").click_input()

        except Exception as e:
            print(e)
            raise e

    def create_excel(self) -> str:
        # Write data to the pre-created sheet (global variable) and make it bold
        ws['A1'] = 'Part Already Exists'
        bold_font = Font(bold=True)
        ws['A1'].font = bold_font

        # Resize column A
        ws.column_dimensions['A'].width = 20

        # Give the sheet a unique name and save it
        file_name = f"output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(file_name)
        print(f"Excel file '{file_name}' created successfully.")

        return file_name

    # endregion File Reading


if __name__ == '__main__':
    Runner = Automator()
    Runner.automate()
